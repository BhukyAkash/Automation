import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

CURRENT_DIR = os.path.dirname(__file__)
EXCEL_PATH = os.path.join(CURRENT_DIR, "STPTestData.xlsx")
BASE_DIR = os.path.join(os.path.dirname(__file__), "..")
DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), "downloads")

# ==== Input: Motor Test Data ======
VEHICLE_CONFIG = {
    "CV": {"reg_col": 1,  "ic_col": 2,  "used_col": 3},   # A, B, C
    "PC": {"reg_col": 5,  "ic_col": 6,  "used_col": 8},   # E, F, H
    "MC": {"reg_col": 10, "ic_col": 11, "used_col": 13},  # J, K, M
}

def get_vehicle_data(vehicle_type):
    cfg = VEHICLE_CONFIG[vehicle_type]
    wb = load_workbook(EXCEL_PATH)
    sheet = wb["Test Data"]

    claimed_row = None

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        used_val = sheet.cell(row=row_idx, column=cfg["used_col"]).value
        used_str = str(used_val).strip().upper() if used_val else ""

        if used_str in ("Y", "RUNNING"):
            continue

        reg   = sheet.cell(row=row_idx, column=cfg["reg_col"]).value
        mykad = sheet.cell(row=row_idx, column=cfg["ic_col"]).value

        if not reg or not mykad:
            break  # Empty row — no more test data

        sheet.cell(row=row_idx, column=cfg["used_col"]).value = "RUNNING"
        wb.save(EXCEL_PATH)
        claimed_row = row_idx
        break

    if claimed_row is None:
        print(f"No test data found for {vehicle_type}")
        return None

    return {
        "vehicle_reg_no": reg,
        "mykad": mykad,
        "claimed_row": claimed_row,
        "vehicle_type": vehicle_type,
    }


def mark_policy_issued(vehicle_type, claimed_row):
    """Call this after policy is successfully issued."""
    cfg = VEHICLE_CONFIG[vehicle_type]
    wb = load_workbook(EXCEL_PATH)
    sheet = wb["Test Data"]
    sheet.cell(row=claimed_row, column=cfg["used_col"]).value = "Y"
    wb.save(EXCEL_PATH)


def reset_on_error(vehicle_type, claimed_row):
    """Call this in except block if policy was NOT issued — returns row to pool."""
    cfg = VEHICLE_CONFIG[vehicle_type]
    wb = load_workbook(EXCEL_PATH)
    sheet = wb["Test Data"]
    sheet.cell(row=claimed_row, column=cfg["used_col"]).value = "N"
    wb.save(EXCEL_PATH)


# ==== Input: PA Test Data ======
def get_pa_data(row=2):
    wb = load_workbook(EXCEL_PATH)
    ws = wb["PA"]

    pa_product     = ws.cell(row=row, column=3).value  # Column C
    occupation_cls = ws.cell(row=row, column=4).value  # Column D
    sum_insured    = ws.cell(row=row, column=5).value  # Column E

    product_map = {
        "pa shield":                "PA Shield",
        "personal accident safe":   "Personal Accident Safe",
        "personal accident shield": "PA Shield",
    }
    if pa_product:
        pa_product = product_map.get(pa_product.strip().lower(), pa_product.strip())

    if isinstance(sum_insured, (int, float)):
        sum_insured = f"{int(sum_insured):,}"
    elif sum_insured:
        sum_insured = str(sum_insured).strip()

    return {
        "pa_product":     pa_product,
        "occupation_cls": occupation_cls.strip() if occupation_cls else None,
        "sum_insured":    sum_insured,
    }


# ==== Internal Helpers ======
def _get_sheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    return wb.create_sheet(sheet_name)

def _next_row(ws):
    row = 2
    while ws.cell(row=row, column=3).value:
        row += 1
    return row

def _serial(ws, row):
    prev = ws.cell(row=row - 1, column=1).value if row > 2 else 0
    return (prev or 0) + 1

def _autofill_prev(ws, row, cols):
    """Copy specified columns from previous row."""
    if row > 2:
        for col in cols:
            prev_val = ws.cell(row=row - 1, column=col).value
            if prev_val:
                ws.cell(row=row, column=col).value = prev_val


# ==== Internal: Motor Writer ======
# A=Serial  B=Auto  C=RV   D=Type     E=Coverage   F=Quote  G=Policy  H=Date
# I=SumInsured  J=ActPrem  K=BasicPrem  L=NCD  M=AfterNCD
# N=GrossPrem   O=SST      P=StampDuty  Q=Total
# R=Auto(18)  S=Auto(19)
def _write_motor(policy_type, coverage_label, quote_number, policy_number,
                sum_insured, act_prem, basic_prem, ncd,
                after_ncd, gross_premium, sst, stamp_duty, total):
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")
    wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
    ws = _get_sheet(wb, "Motor")
    row = _next_row(ws)

    ws.cell(row=row, column=1).value  = _serial(ws, row)
    ws.cell(row=row, column=3).value  = "RV"
    ws.cell(row=row, column=4).value  = policy_type
    ws.cell(row=row, column=5).value  = coverage_label
    ws.cell(row=row, column=6).value  = quote_number
    ws.cell(row=row, column=7).value  = policy_number
    ws.cell(row=row, column=8).value  = datetime.today().strftime("%d-%m-%Y")
    ws.cell(row=row, column=9).value  = sum_insured
    ws.cell(row=row, column=10).value = act_prem
    ws.cell(row=row, column=11).value = basic_prem
    ws.cell(row=row, column=12).value = ncd
    ws.cell(row=row, column=13).value = after_ncd
    ws.cell(row=row, column=14).value = gross_premium
    ws.cell(row=row, column=15).value = sst
    ws.cell(row=row, column=16).value = stamp_duty
    ws.cell(row=row, column=17).value = total

    _autofill_prev(ws, row, cols=[2, 18, 19])  # B, R, S
    wb.save(file_path)
    print("Motor: Quote and Policy numbers saved to Excel")

# ==== Internal: PA Writer ======
# A=Serial  B=Auto  C=NV   D=PA   E=ProductType  F=Quote  G=Policy  H=Date
# I=SumInsured  J=GrossPrem  K=Rebate  L=SST  M=StampDuty  N=Total
# O=Auto(15)  P=Auto(16)
def _write_pa(selected_title, class_ques, quote_number, policy_number,
            sum_insured, gross_premium, rebate, sst, stamp_duty, total):
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")
    wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
    ws = _get_sheet(wb, "PA")
    row = _next_row(ws)

    ws.cell(row=row, column=1).value  = _serial(ws, row)
    ws.cell(row=row, column=3).value  = "NV"
    ws.cell(row=row, column=4).value  = "PA"
    ws.cell(row=row, column=5).value  = selected_title
    ws.cell(row=row, column=6).value  = class_ques
    ws.cell(row=row, column=7).value  = quote_number
    ws.cell(row=row, column=8).value  = policy_number
    ws.cell(row=row, column=9).value  = datetime.today().strftime("%d-%m-%Y")
    ws.cell(row=row, column=10).value  = sum_insured
    ws.cell(row=row, column=11).value = gross_premium  # J
    ws.cell(row=row, column=12).value = rebate         # K
    ws.cell(row=row, column=13).value = sst            # L
    ws.cell(row=row, column=14).value = stamp_duty     # M
    ws.cell(row=row, column=15).value = total          # N

    _autofill_prev(ws, row, cols=[2, 16, 17])  # B, O, P
    wb.save(file_path)
    print("PA: Quote and Policy numbers saved to Excel")

# ==== Internal: Dental Writer ======
# A=Serial  B=Auto  C=NV  D=Dental  E=DentalShield  F=Quote  G=Policy  H=Date
def _write_dental(quote_number, policy_number):
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")
    wb = load_workbook(file_path) if os.path.exists(file_path) else Workbook()
    ws = _get_sheet(wb, "Dental")
    row = _next_row(ws)

    ws.cell(row=row, column=1).value = _serial(ws, row)
    ws.cell(row=row, column=3).value = "NV"
    ws.cell(row=row, column=4).value = "Dental"
    ws.cell(row=row, column=5).value = "Dental Shield"
    ws.cell(row=row, column=6).value = quote_number
    ws.cell(row=row, column=7).value = policy_number
    ws.cell(row=row, column=8).value = datetime.today().strftime("%d-%m-%Y")

    _autofill_prev(ws, row, cols=[2, 15, 16])  # B, O, P
    wb.save(file_path)
    print("Dental: Quote and Policy numbers saved to Excel")

# ==== Public Output Functions ======
def pa_excel(selected_title, class_ques, quote_number, policy_number,
            sum_insured, rebate, gross_premium, sst, stamp_duty, total):
    _write_pa(selected_title, class_ques, quote_number, policy_number,
            sum_insured, rebate, gross_premium, sst, stamp_duty, total)

def dental(quote_number, policy_number):
    _write_dental(quote_number, policy_number)

def mc_excel(selected_coverage, quote_number, policy_number,
            sum_insured, act_prem, basic_prem, ncd,
            after_ncd, gross_premium, sst, stamp_duty, total):
    _write_motor("MC", selected_coverage, quote_number, policy_number,
                sum_insured, act_prem, basic_prem, ncd,
                after_ncd, gross_premium, sst, stamp_duty, total)

def pc_excel(selected_coverage, quote_number, policy_number,
            sum_insured, act_prem, basic_prem, ncd,
            after_ncd, gross_premium, sst, stamp_duty, total):
    _write_motor("PC", selected_coverage, quote_number, policy_number,
                sum_insured, act_prem, basic_prem, ncd,
                after_ncd, gross_premium, sst, stamp_duty, total)

def cv_excel(selected_coverage, quote_number, policy_number,
            sum_insured, act_prem, basic_prem, ncd,
            after_ncd, gross_premium, sst, stamp_duty, total):
    _write_motor("CV", selected_coverage, quote_number, policy_number,
                sum_insured, act_prem, basic_prem, ncd,
                after_ncd, gross_premium, sst, stamp_duty, total)