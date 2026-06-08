import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import Workbook, load_workbook
from excel_utils import get_vehicle_data
from datetime import datetime
from base_login import incep_date, login, navigation, pc_moto, issue_policy
from extension import mc_extension
from popup_utils import ask_popup
from test_mail import send_email

# ---- Path References ----
BASE_DIR = os.path.join(os.path.dirname(__file__), "..")               # D:\Automation\pages
DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), "downloads")  # D:\Automation\pages\NB\downloads

def test_mc_motor(page):

    try:
        print("====================== Issuance of MC policy ==================")
        login(page)
        navigation(page)
        pc_moto(page)

        # ========= FIRST SCREEN ===========
        
        # ---- VEHICLE REG ----
        vehicle_data = get_vehicle_data("MC")
        page.get_by_role("textbox").first.fill(vehicle_data["vehicle_reg_no"])

        #---- Place of Use ----
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Johor").click()

        # ---- Vehicle Search ----
        page.get_by_role("button", name="search Vehicle Search").click()
        page.wait_for_timeout(2000)

        # --- Engine Capacity field ----
        cc_input = page.locator('input#cc')
        if cc_input.is_visible():
            current_value = cc_input.input_value().strip()
            if current_value == "" or current_value == "0":
                cc_input.dblclick()
                cc_input.fill("1200")
            else:
                print(f"Engine Capacity: {current_value}")

        # --- Seating Capacity field ----
        seat_input = page.locator('input#seatCapacity')

        if seat_input.is_visible():
            current_value = seat_input.input_value().strip()
            if current_value == "" or current_value == "0":
                seat_input.dblclick()
                seat_input.fill("2")
            else:
                print(f"Seating Capacity: {current_value}")

        # ---- Vehicle Age from input (Screen 1) ----
        vehicle_age_text = ""
        try:
            vehicle_age_locator = page.locator("mat-form-field").filter(has_text="Vehicle Age").locator("#vehicleAge")
            vehicle_age_text = vehicle_age_locator.input_value().strip()
        except:
            pass

        # ---- SAVE VEHICLE DETAILS ----
        search_vehicle = page.get_by_role("button", name="Save Vehicle Info").first
        try:
            search_vehicle.wait_for(state="visible", timeout=2000)
            search_vehicle.click()
        except:
            print("Save Vehicle Info button not available")

        # ---- Vehicle Age from span (Screen 2 fallback) ----
        if not vehicle_age_text:    
            try:
                page.locator("span.status-text").first.wait_for(state="visible", timeout=10000)
                spans = page.locator("span.status-text").all()
                vehicle_age_text = spans[6].inner_text().strip()
            except:
                vehicle_age_text = "0"
                print("Vehicle Age not found, defaulting to 0")

        vehicle_age = int(vehicle_age_text) if vehicle_age_text else 0
        print(f"Vehicle Age: {vehicle_age} years")

        
        # ========== SECOND SCREEN ==========

        # ---- COVERAGE TYPE (read default) ----
        selected_coverage = page.locator("#mat-select-value-9 span.mat-select-min-line").inner_text().strip()
        print("Default Coverage type: ", selected_coverage)

        # ---- COVERAGE TYPE (change only if default is not Third Party) ----
        if selected_coverage != "Third Party":
            page.locator("#mat-select-value-9").click()
            if vehicle_age >= 15:
                page.get_by_role("option", name="Third Party").click()
            else:
                page.get_by_role("option", name="Comprehensive").click()

        selected_coverage = page.locator("#mat-select-value-9 span.mat-select-min-line").inner_text().strip()
        print("Selected Coverage type: ", selected_coverage)

        # ---- COVERAGE DATE -----
        incep_date(page)

        # ---- MARKET VALUE ----
        market_value_text = page.locator("mat-form-field").filter(has_text="Market Value").locator("#ismMarketValue").input_value().strip()
        market_value = int(float(market_value_text.replace(",", "")))
        print(f"Market Value: {market_value}")

        # ---- VEHICLE SUM INSURED ----
        if selected_coverage != "Third Party":
            sum_insured = str(market_value) if market_value > 1000 else "1000"
            print(f"Sum Insured: {sum_insured}")

            page.locator("dx-input-currency").filter(has_text="* Vehicle Sum Insured *").locator("#sumInsured").click()
            page.locator("dx-input-currency").filter(has_text="* Vehicle Sum Insured *").locator("#sumInsured").fill(sum_insured)
        else:
            print("Third Party selected, skipping Sum Insured")

        #---- MYKAD ID ----
        page.locator("mat-form-field").filter(has_text="ID # * help").locator("#id").fill(vehicle_data["mykad"])

        #----NAME OF THE PH ----
        page.locator("mat-form-field").filter(has_text="Name as per ID *").locator("#legalName").fill("Motor Cycle")
        page.get_by_role("button", name="search Validate Owner as per").click()

        # ==== Multi Contract / Extensions ====
        print("======== Extension Coverage Selection ========")

        answer = ask_popup(
            question="Do you want to explore Extensions screen?",
            title="Extension Coverage Selection",
        )

        if answer == "yes":
            mc_extension(page, selected_coverage)
            print("Extensions added successfully")
        else:
            print("No Extensions Selected")

        page.pause()
        # ---- NCD value ----
        page.wait_for_timeout(4000)
        ncd_value = page.locator("#currentNCD input.mat-input-element").input_value()
        print("NCD Value:", ncd_value)

        #---- SAVE & NEXT BUTTON -----
        page.get_by_role("button", name="Save & Next").click()
        print("Registration Number is Triggered to ISM")


        # ========== THIRD SCREEN === COVER DETAILS ==========

        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Less than 2 years").click()

        # ---- CHECK IF YES BUTTON EXISTS AND IS ENABLED ----
        try:
            yes_button = page.get_by_role("button", name="Yes").first
            yes_button.wait_for(state="visible", timeout=5000)
            yes_button.click()
            page.wait_for_timeout(1000)
            print("Yes button clicked")
        except:
            print("Yes button not visible, skipping")

        # ---- CHECK IF ADDRESS ALREADY EXISTS ----
        try:
            add_button = page.locator("button[name='Add'], button:has-text('Add')").first
            add_button.wait_for(state="visible", timeout=5000)
            add_button.click()
            print("Add button clicked")
            page.wait_for_timeout(2000)
        except:
            print("Add button not visible, skipping")

        # ---- STATE ---- (runs for both cases)
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Johor").click()
        page.wait_for_timeout(3000)

        # ---- PINCODE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="81100").click()
        page.wait_for_timeout(2000)

        # ---- STREET ADDRESS ----
        page.get_by_role("combobox", name="Address Line").click()
        page.get_by_role("option", name="Taman Desa Harmoni", exact=True).click()
        page.wait_for_timeout(2000)

        # ---- SAVE BUTTON (if address is added) ----
        address_save = page.locator("button#save")
        if address_save.is_visible():
            address_save.click()

        # Locate the element that contains the quote reference
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        quote_number = quote_text.strip()
        print("Quote Number:", quote_number)

        # ---- DECLARATION STATEMENTS ----
        page.get_by_text("We respect your privacy and").click()
        page.get_by_text("I hereby confirm that I have").click()

        # ---- GENERATE & DOWNLOAD QUOTE -----
        page.get_by_role("button", name="Generate Quote").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as(os.path.join(DOWNLOADS_DIR, "MC_quote.pdf"))
        print("Quote PDF downloaded successfully.")

        # ==== Issue Policy function ====
        policy_number = issue_policy(page)
        
        # ---- Download the policy schedule ----
        page.get_by_role("button", name="Download & e-mail Policy").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as(os.path.join(DOWNLOADS_DIR, "MC_policy.pdf"))
        
        print("Policy is Issued and Schedule letter downloaded successfully.")

        # --------- SAVE TO EXCEL ---------
        
        file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

        # Load or create workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # ---- Find next empty row based on Column C (NV/RV) ----
        row = 2  # start after header
        while ws.cell(row=row, column=3).value:
            row += 1

        # ---- Serial Number in Column A ----
        if row == 2:
            serial_no = 1
        else:
            prev_serial = ws.cell(row=row - 1, column=1).value
            serial_no = (prev_serial or 0) + 1

        # ---- Policy Type ----
        registration = "RV"
        policy_type = "MC"

        inception_date_excel = datetime.today().strftime("%d-%m-%Y")

        # ---- Write data ----
        ws.cell(row=row, column=1).value = serial_no             # Column A - Serial Number
        ws.cell(row=row, column=3).value = registration          # Column C - NV/RV
        ws.cell(row=row, column=4).value = policy_type           # Column D - Policy Type
        ws.cell(row=row, column=5).value = selected_coverage     # Column E - Coverage Type
        ws.cell(row=row, column=6).value = quote_number          # Column F - Quote Number
        ws.cell(row=row, column=7).value = policy_number         # Column G - Policy Number
        ws.cell(row=row, column=8).value = inception_date_excel  # Column H

        # ---- Auto-fill Column I & J from previous row (like Ctrl+D) ----
        if row > 2:
            prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
            prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
            prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

            if prev_col_b:
                ws.cell(row=row, column=2).value = prev_col_b   # Column B
            if prev_col_i:
                ws.cell(row=row, column=9).value = prev_col_i   # Column I
            if prev_col_j:
                ws.cell(row=row, column=10).value = prev_col_j  # Column J

        # ---- Save file ----
        wb.save(file_path)

        print("In Excel, Quote and Policy numbers captured successfully")

        # -------- SEND EMAIL ---------
        try:
            send_email()
        except Exception as e:
            print("Email failed:", e)

    finally:
        page.bring_to_front()
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()
        page.wait_for_timeout(15000)