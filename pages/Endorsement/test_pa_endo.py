import os
import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from tips_endo import endorsement
from openpyxl import load_workbook


# Run Command: pytest -s endorsement\test_pa_endo.py --product=pa
def test_pa(page, request):
    try:
        policy_number = endorsement(page, request)

        # ---- Endorsement Reason ----
        reason = "Correct Insured Person Details"
        page.get_by_text(reason).click()
        print(f"Endorsement Reason: {reason}")

        # --- Sub Reason Drop down Selection ---
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Occupation").click()

        # --- Occupation Class ---
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Class 1").click()

        # --- Declaration Statement ---
        page.get_by_text("I/We declare that the Policy Information Provided").click()

        # --- Save & Next ----
        page.get_by_role("button", name="Save").click()
        page.get_by_role("button", name="Next").click()

        # --- Endo Quote Reference Number ---
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        endo_quote = quote_text.strip()
        print("Endorsement Quote Number:", endo_quote)

        page.get_by_role("button", name="Proceed").click()
        page.get_by_role("button", name="Submit for Processing").click()

        print("Endorsement performed successfully")


        # ====== Save to Excel ======
        BASE_DIR = os.path.join(os.path.dirname(__file__), "..")  # D:\Automation\pages
        file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active

            matched_row = None
            for row in ws.iter_rows(min_row=2):
                cell_policy = row[6]  # Column G (0-indexed = 6) = Policy Number
                if cell_policy.value and str(cell_policy.value).strip() == str(policy_number).strip():
                    matched_row = cell_policy.row
                    break

            ws.cell(row=matched_row, column=11).value = endo_quote  # Column K
            wb.save(file_path)
            print(f"Endo Quote Reference '{endo_quote}' written into Excel, Row {matched_row} (Policy: {policy_number})")

    finally:
        page.bring_to_front()
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()
        page.wait_for_timeout(15000)