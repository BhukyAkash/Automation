import os
import random
from openpyxl import Workbook, load_workbook
from base_login import login, navi_pa
from conftest import page
from test_mail import send_email
from datetime import datetime, timedelta

# ---------- UTIL FUNCTIONS ----------

def log(step, message):
    print(f"[{step}] {message}")

def generate_mykad():
    d = datetime(1970,1,1) + timedelta(days=random.randint(0, (datetime.today()-datetime(1950,1,1)).days))
    return f"{d:%y%m%d}-{random.randint(10,99)}-{random.randint(1000,9999)}"

def safe_click(locator, retries=3):
    for i in range(retries):
        try:
            locator.click()
            return
        except:
            print(f"Retry {i+1}...")
    raise Exception("Click failed after retries")

# ---------- MAIN TEST ----------

def test_PA(page):

    try:
        # ---- LOGIN ----
        login(page)
        navi_pa(page)
        log("LOGIN", "User logged in & navigated")

        # ========= FIRST SCREEN ===========

        # ---- MYKAD ID ----
        page.locator("#dx-input-0").nth(1).fill(generate_mykad())

        # ---- PH NAME ----
        page.locator("#dx-input-1").nth(1).fill("PERSONAL ACCIDENT")

        # ---- CLASSIFICATION ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Class 2").click()

        # ---- PRODUCT ----
        page.locator("[formcontrolname='paProducts']").click()
        page.get_by_role("option", name="Personal Accident Safe").click()

        # ---- PLAN TYPE ----
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="100,000").click()

        # ---- WEEKLY BENEFIT ----
        page.locator("mat-radio-button:has-text('No')").nth(1).click()
        log("STEP", "Weekly Benefit not selected")

        page.locator("#dx-checkbox-1").click()

        # ---- SAVE & NEXT ----
        safe_click(page.get_by_role("button", name="Save & Next"))
        page.wait_for_selector("text=Address", timeout=15000)

        # ========== SECOND SCREEN ==========

        add_button = page.get_by_role("button", name="Add").first

        if add_button.is_visible(timeout=3000):
            log("ADDRESS", "Adding new address")
            add_button.click()

        # ---- STATE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Johor").click()

        # ---- PINCODE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="81100").click()

        # ---- ADDRESS ----
        page.get_by_role("combobox", name="Address Line").click()
        page.get_by_role("option", name="Taman Desa Harmoni", exact=True).click()

        # ---- CONTACT ----
        page.get_by_role("textbox", name="123456789").fill("123456789")
        page.get_by_role("textbox", name="Enter").fill("akash@serole.com")

        # ---- UNDERWRITING ----
        radios = page.get_by_role("radio", name="No")
        for i in range(5):
            radios.nth(i).check()

        # ---- DECLARATIONS ----
        page.get_by_text("1. Are any of the Insured").click()
        page.get_by_text("We respect your privacy").click()
        page.get_by_text("I hereby confirm that I have").click()

        # ---- GET QUOTE NUMBER ----
        page.wait_for_selector("text=Quote Reference #")
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        quote_number = quote_text.strip()
        log("QUOTE", quote_number)

        # ---- GENERATE QUOTE ----
        safe_click(page.get_by_role("button", name="Generate Quote"))

        if page.locator("text=Download Quote & PDS Documents").is_visible():
            log("QUOTE", "Generated successfully")
        else:
            raise Exception("Quote generation failed")

        # ---- DOWNLOAD QUOTE ----
        safe_click(page.get_by_role("button", name="Download Quote & PDS Documents"))

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Download").click()

        download = download_info.value
        download.save_as(f"downloads/PA_quote_{quote_number}.pdf")

        # ---- ISSUE POLICY ----
        safe_click(page.get_by_role("button", name="Issue Policy"))
        safe_click(page.get_by_role("button", name="Accept & Proceed"))
        log("POLICY", "Issued successfully")

        page.wait_for_timeout(10000)
        page.reload()

        # ---- DOWNLOAD POLICY ----
        safe_click(page.get_by_role("button", name="Download & e-mail Policy"))
        page.get_by_text("Download Policy Schedule").click()

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Download").click()

        download = download_info.value

        # ---- GET POLICY NUMBER ----
        policy_locator = page.locator("text=Policy #:")
        policy_locator.wait_for()
        policy_text = policy_locator.text_content()
        policy_number = policy_text.replace("Policy #:", "").strip()

        download.save_as(f"downloads/PA_policy_{policy_number}.pdf")

        log("POLICY", policy_number)

        # ---- VALIDATION ----
        if not quote_number or not policy_number:
            raise Exception("Invalid data for Excel")

        # ---- SAVE TO EXCEL ----
        file_path = "UATStability.xlsx"

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        row = 2
        while ws.cell(row=row, column=4).value:
            row += 1

        ws.cell(row=row, column=4).value = "PA"
        ws.cell(row=row, column=6).value = quote_number
        ws.cell(row=row, column=7).value = policy_number
        ws.cell(row=row, column=8).value = datetime.today().strftime("%d-%m-%Y")

        wb.save(file_path)

        log("EXCEL", "Data saved successfully")

        # ---- EMAIL ----
        send_email()
        log("EMAIL", "Sent successfully")

    except Exception as e:
        page.screenshot(path="error.png")
        print("ERROR:", e)

    finally:
        page.wait_for_timeout(5000)
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()