import os
from openpyxl import Workbook, load_workbook
from excel_utils import get_vehicle_data
from datetime import datetime
from base_login import login, navigation, pc_moto
from test_mail import send_email

def test_mc_motor(page):

    try:
        login(page)
        navigation(page)
        pc_moto(page)

        # ========= FIRST SCREEN ===========
        
        # ---- VEHICLE REG ----
        vehicle_data = get_vehicle_data("MC")
        page.get_by_role("textbox").first.fill(vehicle_data["vehicle_reg_no"])

        #---- Place of Use ----
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Johor").click()

        #---- Vehicle Search ----
        page.get_by_role("button", name="search Vehicle Search").click()

        page.wait_for_timeout(5000)

        #---- MYKAD ID ----
        page.locator("mat-form-field").filter(has_text="ID # * help").locator("#id").fill(vehicle_data["mykad"])

        #----NAME OF THE PH ----
        page.locator("mat-form-field").filter(has_text="Name as per ID *").locator("#legalName").fill("Motor Cycle")
        page.get_by_role("button", name="search Validate Owner as per").click()

        # ---- NCD value ----
        page.wait_for_timeout(4000)
        ncd_value = page.locator("#currentNCD input.mat-input-element").input_value()
        print("NCD Value:", ncd_value)

        #---- SAVE & NEXT BUTTON -----
        page.get_by_role("button", name="Save & Next").click()
        print("Registration Number is Triggered to ISM")


        # ========== THIRD SCREEN === COVER DETAILS ==========

        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Less than 2 years").click()

        # ---- CHECK IF YES BUTTON EXISTS AND IS ENABLED ----
        yes_button = page.get_by_role("button", name="Yes").first

        if yes_button.is_visible():
                yes_button.click()
                page.wait_for_timeout(1000)

        # ---- CHECK IF ADDRESS ALREADY EXISTS ----
        add_button = page.locator("button[name='Add'], button:has-text('Add')").first
        if add_button.is_visible():
            add_button.click()
            page.wait_for_timeout(1000)

        # ---- STATE ---- (runs for both cases)
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Johor").click()
        page.wait_for_timeout(5000)

        # ---- PINCODE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="81100").click()
        page.wait_for_timeout(2000)

        # ---- STREET ADDRESS ----
        page.get_by_role("combobox", name="Address Line").click()
        page.get_by_role("option", name="Taman Desa Harmoni", exact=True).click()
        page.wait_for_timeout(2000)

        # ---- SAVE BUTTON (if address is added) ----
        address_save = page.locator("button#save")
        if address_save.is_visible():
            address_save.click()
            page.locator("//label[@for='3']//div[@class='box-card justify-content-between']").wait_for(state="visible")
            page.locator("//label[@for='3']//div[@class='box-card justify-content-between']").click()

        # Locate the element that contains the quote reference
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        quote_number = quote_text.strip()
        print("Quote Number:", quote_number)

        # ---- DECLARATION STATEMENTS ----
        page.get_by_text("We respect your privacy and").click()
        page.get_by_text("I hereby confirm that I have").click()

        # ---- GENERATE & DOWNLOAD QUOTE -----
        page.get_by_role("button", name="Generate Quote").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as("downloads/MC_quote.pdf")
        print("Quote PDF downloaded successfully.")

        # ----- PROCEED TO POLICY ISSUANCE ----
        page.get_by_role("button", name="Proceed to Policy Issuance").click()

        # ------- POLICY ISSUANCE -------
        page.get_by_role("button", name="Issue Policy").click()

        page.wait_for_timeout(30000)

        page.reload()

        # ---- Printing the policy number ---
        policy_locator = page.locator("text=Policy #:")
        policy_locator.wait_for()
        policy_text = policy_locator.text_content()
        policy_number = policy_text.replace("Policy #:", "").strip()
        print("Policy Number:", policy_number)


        # ---- Download the policy schedule ----
        page.get_by_role("button", name="Download & e-mail Policy").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as("downloads/MC_policy.pdf")
        print("Policy is Issued and Schedule letter downloaded successfully.")

        # --------- SAVE TO EXCEL ---------
        
        file_path = "UATStability.xlsx"

        # Load or create workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # ---- Find next empty row based on Column D (Motor Type) ----
        row = 2  # start after header
        while ws.cell(row=row, column=4).value:
            row += 1

        # ---- Policy Type ----
        policy_type = "MC"

        inception_date_excel = datetime.today().strftime("%d-%m-%Y")

        # ---- Write data ----
        ws.cell(row=row, column=4).value = policy_type
        ws.cell(row=row, column=6).value = quote_number
        ws.cell(row=row, column=7).value = policy_number
        ws.cell(row=row, column=8).value = inception_date_excel

        # ---- Save file ----
        wb.save(file_path)

        # -------- SEND EMAIL ---------
        send_email()

    finally:
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()
        page.wait_for_timeout(15000)