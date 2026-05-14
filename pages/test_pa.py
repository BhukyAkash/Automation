import os
from base_login import login, navi_pa
from test_mykad_id import generate_mykad
from datetime import datetime
from openpyxl import Workbook, load_workbook
from test_mail import send_email

def test_PA(page):

    try:
        login(page)
        navi_pa(page)
        print("User successfully logged in and navigated to Personal Accident section")

        # ========= FIRST SCREEN ===========

        # ---- MYKAD ID ----
        page.locator("#dx-input-0").nth(1).click()
        mykad = generate_mykad()
        page.locator("#dx-input-0").nth(1).fill(mykad)
        print("MyKad ID: ", mykad)

        # ---- PH NAME ----
        page.locator("#dx-input-1").nth(1).click()
        page.locator("#dx-input-1").nth(1).fill("PERSONAL ACCIDENT")

        # ---- INTERNAL CLASSIFICATION ----
        #page.locator("[formcontrolname='internalClassification']").click()
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Class 2").click()

        # ---- PRODUCT SELECTION ----
        selected_title = "Personal Accident Safe"    
        #selected_title = "PA Shield"
        page.locator("[formcontrolname='paProducts']").click()
        page.get_by_role("option", name=selected_title).click()
        print("Selected Product:", selected_title)

        # ---- PLAN TYPE ----
        page.locator("#occupation-description").filter(has_text="Sum Insured").click()
        page.get_by_role("option", name="100,000").click()

        # ---- WEEKLY BENEFIT ----
        #page.locator("#mat-radio-9 > .mat-radio-label > .mat-radio-container > .mat-radio-outer-circle").click()
        page.locator("mat-radio-button:has-text('No')").nth(1).click()
        print("Weekly Benefit not selected")

        page.locator("#dx-checkbox-1 > .mat-checkbox-layout > .mat-checkbox-inner-container").click()

        # ---- SAVE & NEXT ----
        page.get_by_role("button", name="Save & Next").click()
        page.wait_for_timeout(10000)

        # ========== SECOND SCREEN ==========


        # ========== SCREEN 2 DEBUG ==========
        page.wait_for_timeout(3000)

        '''print("=== SCREEN 2 DEBUG ===")
        print("URL:", page.url)
        print("All buttons:")
        for btn in page.get_by_role("button").all():
            print(f"  BUTTON: '{btn.inner_text().strip()}'")

        print("All mat-select (formcontrolname):")
        for sel in page.locator("mat-select").all():
            print(f"  MAT-SELECT formcontrolname: '{sel.get_attribute('formcontrolname')}'")

        print("mat-select-placeholder count:", page.locator(".mat-select-placeholder").count())
        print("Yes visible:", page.get_by_role("button", name="Yes").first.is_visible())
        print("Add visible:", page.get_by_role("button", name="Add").first.is_visible())
        print("=== END SCREEN 2 DEBUG ===")'''

        # ---- CHECK IF YES BUTTON EXISTS AND IS ENABLED ----
        try:
            page.get_by_role("button", name="Yes").first.wait_for(state="visible", timeout=5000)
            page.get_by_role("button", name="Yes").first.click()
            page.wait_for_timeout(2000)
        except:
            pass

        # ---- CHECK IF ADDRESS ALREADY EXISTS ----
        add_button = page.locator("button[name='Add'], button:has-text('Add')").first
        try:
            add_button.wait_for(state="visible", timeout=3000)
            add_button.click()
            page.wait_for_timeout(1000)
        except:
            pass  

        # ---- STATE ---- (runs for both cases)
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Johor").click()
        page.wait_for_timeout(2000)

        # ---- PINCODE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="81100").click()
        page.wait_for_timeout(1000)

        # ---- STREET ADDRESS ----
        page.get_by_role("combobox", name="Address Line").click()
        page.get_by_role("option", name="Taman Desa Harmoni", exact=True).click()
        page.wait_for_timeout(1000)

        # ---- SAVE BUTTON (if address is added) ----
        address_save = page.get_by_role("button", name="Save")
        if address_save.is_visible():
            address_save.click()            


        # ---- CONTACT DETAILS ----
        page.get_by_role("textbox", name="123456789").fill("123456789")
        page.get_by_role("textbox", name="Enter").fill("akash@serole.com")


        # ---- UNDERWRITING QUESTIONS ----
        radios = page.get_by_role("radio", name="No")
        for i in range(5):
            radios.nth(i).check()

        # ---- Declaration Statements ----
        page.get_by_text("We respect your privacy").click()
        page.get_by_text("I hereby confirm that I have").click()

        # ---- Get Quote Number ----
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        quote_number = quote_text.strip()
        print("Quote Number:", quote_number)
        
        # ---- GENERATE & DOWNLOAD QUOTE -----
        page.get_by_role("button", name="Generate Quote").wait_for()
        page.get_by_role("button", name="Generate Quote").click()
    
        page.get_by_role("button", name="Download Quote & PDS Documents").click()
        page.locator("form").get_by_text("Download Quote & PDS Documents").click()

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Download").click()
        download = download_info.value
        download.save_as("downloads/PA_quote.pdf")
        page.get_by_text("close", exact=True).click()
        print("Quote PDF Generated successfully")

        page.wait_for_timeout(10000)

        # ---- ISSUE POLICY & DOWNLOAD POLICY SCHEDULE ----
        page.get_by_role("button", name="Issue Policy").click()
        page.get_by_role("button", name="Accept & Proceed").click()
        
        page.wait_for_timeout(25000)
        page.reload()

        # ---- Download the policy schedule ----
        page.get_by_role("button", name="Download & Email Policy").click()
        page.get_by_text("Download Policy Schedule").click()
        print("Policy Issued successfully")

        page.wait_for_timeout(10000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Download").click()
        download = download_info.value
        download.save_as("downloads/PA_policy.pdf")
        page.get_by_text("close", exact=True).click()

        # ---- Printing the policy number ----
        policy_locator = page.locator("text=Policy #:")
        policy_locator.wait_for()
        policy_text = policy_locator.text_content()
        policy_number = policy_text.replace("Policy #:", "").strip()
        print("Policy Number:", policy_number)

        print("Policy Schedule downloaded successfully")

        # ------- SAVE TO EXCEL -------

        file_path = "UATStability.xlsx"

        # Load or create workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # ---- Find next empty row based on Column C (NV/RV) ----
        row = 2
        while ws.cell(row=row, column=3).value:
            row += 1

        # ---- Serial Number in Column A ----
        if row == 2:
            serial_no = 1
        else:
            prev_serial = ws.cell(row=row - 1, column=1).value
            serial_no = (prev_serial or 0) + 1

        # ---- Policy Type ----
        registration = "NV"
        policy_type = "PA"
        inception_date_excel = datetime.today().strftime("%d-%m-%Y")

        # ---- Write data ----
        ws.cell(row=row, column=1).value = serial_no        # Column A - Serial Number
        ws.cell(row=row, column=3).value = registration     # Column C - NV/RV
        ws.cell(row=row, column=4).value = policy_type      # Column D - Policy Type
        ws.cell(row=row, column=5).value = selected_title   # Column E - Coverage Type
        ws.cell(row=row, column=6).value = quote_number     # Column F - Quote Number
        ws.cell(row=row, column=7).value = policy_number    # Column G - Policy Number
        ws.cell(row=row, column=8).value = inception_date_excel

        # ---- Auto-fill Column B, I & J from previous row (like Ctrl+D) ----
        if row > 2:
            prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
            prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
            prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

            if prev_col_b:
                ws.cell(row=row, column=2).value = prev_col_b   # Column B
            if prev_col_i:
                ws.cell(row=row, column=9).value = prev_col_i   # Column I
            if prev_col_j:
                ws.cell(row=row, column=10).value = prev_col_j  # Column J

        # ---- Save file ----
        wb.save(file_path)

        print("In Excel, Quote and Policy numbers captured successfully")


        # ---- SEND EMAIL ----
        try:
            send_email()
        except Exception as e:
            print("Email failed:", e)

    finally:
        page.wait_for_timeout(15000)
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()
        page.wait_for_timeout(7000)