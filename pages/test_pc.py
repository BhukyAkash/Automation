import os
from base_login import login, navigation, pc_moto
from excel_utils import get_vehicle_data
from datetime import datetime
from openpyxl import Workbook, load_workbook
from test_mail import send_email


def test_pc_motor(page):

    try:
        login(page)
        navigation(page)
        pc_moto(page)

        # ========= FIRST SCREEN ===========
        
        # ---- VEHICLE REG ----
        vehicle_data = get_vehicle_data("PC")
        page.get_by_role("textbox").first.fill(vehicle_data["vehicle_reg_no"])

        # ---- Place of Use ----
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Johor").click()

        # ---- Vehicle Search ----
        page.get_by_role("button", name="search Vehicle Search").click()
        page.wait_for_timeout(5000)
        
        # ---- Vehicle Age (to determine coverage type) ---- 
        vehicle_age_locator = page.locator("mat-form-field").filter(has_text="Vehicle Age").locator("#vehicleAge")
        vehicle_age_text = vehicle_age_locator.input_value().strip()

        vehicle_age = int(vehicle_age_text)
        print(f"Vehicle Age: {vehicle_age} years")

        # ---- SAVE VEHICLE DETAILS  ----
        search_vehicle = page.get_by_role("button", name="Save Vehicle Info").first
        try:
            search_vehicle.wait_for(state="visible", timeout=5000)
            search_vehicle.click()
            page.wait_for_load_state("networkidle")
        except:
            print("Save Vehicle Info button not available")

        # ========== SECOND SCREEN ==========

        # ---- COVERAGE TYPE ----
        page.locator("#mat-select-value-9").click()
        if vehicle_age >= 20:
            page.get_by_role("option", name="TP, Fire & Theft").click()
        else:
            page.get_by_role("option", name="Comprehensive").click()

        selected_coverage = page.locator("#mat-select-value-9").inner_text().strip()
        print("Selected Coverage type: ", selected_coverage)

        # ---- COVERAGE DATE -----
            # today date
        today = datetime.today()

            # Angular Material aria-label format
        aria_date = today.strftime("%B %d, %Y").replace(" 0", " ")

            # Open calendar
        page.locator("mat-form-field") \
            .filter(has_text="Inception Date * event") \
            .get_by_label("Open calendar") \
            .click()

            # Select today
        page.get_by_role("gridcell", name=aria_date).click()

        # ---- MARKET VALUE ----
        market_value_text = page.locator("mat-form-field").filter(has_text="Market Value").locator("#ismMarketValue").input_value().strip()
        market_value = int(float(market_value_text.replace(",", "")))
        print(f"Market Value: {market_value}")

        # ---- VEHICLE SUM INSURED ----
        sum_insured = str(market_value) if market_value > 5000 else "5000"
        print(f"Sum Insured: {sum_insured}")

        page.locator("dx-input-currency").filter(has_text="* Vehicle Sum Insured *").locator("#sumInsured").click()
        page.locator("dx-input-currency").filter(has_text="* Vehicle Sum Insured *").locator("#sumInsured").fill(sum_insured)

        #---- MYKAD ID ----
        page.locator("mat-form-field").filter(has_text="ID # * help").locator("#id").fill(vehicle_data["mykad"])

        #----NAME OF THE PH ----
        page.locator("mat-form-field").filter(has_text="Name as per ID *").locator("#legalName").fill("PC")
        page.get_by_role("button", name="search Validate Owner as per").click()
        print("Regio's validated successfully")

        # ---- NCD value ----
        page.wait_for_timeout(7000)
        ncd_value = page.locator("#currentNCD input.mat-input-element").input_value()
        print("NCD Value:", ncd_value)

        #---- SAVE & NEXT BUTTON -----
        page.get_by_role("button", name="Save & Next").click()


        # ========== THIRD SCREEN === COVER DETAILS ==========

        # ---- DRIVER EXPERIENCE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Less than 2 years").click()

        # ---- CHECK IF YES BUTTON EXISTS AND IS ENABLED ----
        yes_button = page.get_by_role("button", name="Yes").first

        if yes_button.is_visible():
                yes_button.click()
                page.wait_for_timeout(1000)

        # ---- CHECK IF ADDRESS ALREADY EXISTS ----
        add_button = page.locator("button[name='Add'], button:has-text('Add')").first
        if add_button.is_visible():
            add_button.click()
            page.wait_for_timeout(1000)

        # ---- STATE ---- (runs for both cases)
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Johor").click()
        page.wait_for_timeout(3000)

        # ---- PINCODE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="81100").click()
        page.wait_for_timeout(2000)

        # ---- STREET ADDRESS ----
        page.get_by_role("combobox", name="Address Line").click()
        page.get_by_role("option", name="Taman Desa Harmoni", exact=True).click()
        page.wait_for_timeout(2000)

        # ---- SAVE BUTTON (if address is added) ----
        address_save = page.locator("button#save")
        if address_save.is_visible():
            address_save.click()
        
        # ---- Garage Types ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Public Road").click()
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="No Alarm(WITHOUT MECHANICAL").click()
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Driver’s Side Airbags (1)").click()

        # Locate the element that contains the quote reference
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        quote_number = quote_text.strip()
        print("Quote Number:", quote_number)

        # ---- DECLARATION STATEMENTS ----
        page.get_by_text("We respect your privacy and").click()
        page.get_by_text("I hereby confirm that I have").click()

        # ---- GENERATE & DOWNLOAD QUOTE -----
        page.get_by_role("button", name="Generate Quote").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as("downloads/PC_quote.pdf")
        print("Quote PDF downloaded successfully.")

        # ----- PROCEED TO POLICY ISSUANCE ----
        page.get_by_role("button", name="Proceed to Policy Issuance").click()

        # ------- POLICY ISSUANCE -------
        page.get_by_role("button", name="Issue Policy").click()

        page.wait_for_timeout(30000)

        page.reload()

        # ---- Download the policy schedule ----
        page.get_by_role("button", name="Download & e-mail Policy").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as("downloads/PC_policy.pdf")
        

        # ---- Printing the policy number ---
        policy_locator = page.locator("text=Policy #:")
        policy_locator.wait_for()
        policy_text = policy_locator.text_content()
        policy_number = policy_text.replace("Policy #:", "").strip()
        print("Policy Number:", policy_number)
        
        print("Policy is Issued and Schedule letter downloaded successfully.")

        # --------- SAVE TO EXCEL ---------
        
        file_path = "UATStability.xlsx"

        # Load or create workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # ---- Find next empty row based on Column C (NV/RV) ----
        row = 2  
        while ws.cell(row=row, column=3).value:
            row += 1
        
        # ---- Serial Number in Column A ----
        if row == 2:
            serial_no = 1
        else:
            prev_serial = ws.cell(row=row - 1, column=1).value
            serial_no = (prev_serial or 0) + 1

        # ---- Policy Type ----
        registration = "RV"
        policy_type = "PC"

        inception_date_excel = datetime.today().strftime("%d-%m-%Y")

        # ---- Write data ----
        ws.cell(row=row, column=1).value = serial_no         # Column A - Serial Number
        ws.cell(row=row, column=3).value = registration      # Column C - NV/RV
        ws.cell(row=row, column=4).value = policy_type       # Column D - Policy Type
        ws.cell(row=row, column=5).value = selected_coverage # Column E - Coverage Type
        ws.cell(row=row, column=6).value = quote_number      # Column F - Quote Number
        ws.cell(row=row, column=7).value = policy_number     # Column G - Policy Number
        ws.cell(row=row, column=8).value = inception_date_excel       

        # ---- Auto-fill Column I & J from previous row (like Ctrl+D) ----
        if row > 2: 
            prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
            prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
            prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

            if prev_col_b:
                ws.cell(row=row, column=2).value = prev_col_b   # Column B
            if prev_col_i:
                ws.cell(row=row, column=9).value = prev_col_i   # Column I
            if prev_col_j:
                ws.cell(row=row, column=10).value = prev_col_j  # Column J

        # ---- Save file ----
        wb.save(file_path)

        print("In Excel, Quote and Policy numbers captured successfully")

        # -------- SEND EMAIL ---------
        try:
            send_email()
        except Exception as e:
            print("Email failed:", e)


    finally:
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()
        page.wait_for_timeout(15000)