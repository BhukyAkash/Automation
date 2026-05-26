from datetime import datetime
from openpyxl import load_workbook
import os

# --- PA Class Change ---
def class_change(page):
        # ---- Endorsement Reason ----
        reason = "Correct Insured Person Details"
        page.get_by_text(reason).click()
        print(f"Endorsement Reason: {reason}")

        # --- Sub Reason Drop down Selection ---
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Occupation").click()

        # --- Occupation Class ---
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Class 1").click()

# --- PA Sum Insured Change ---
def SI_change(page):
        # ---- Endorsement Reason ----
        reason = "Amend Coverage / Extensions"
        page.get_by_text(reason).click()
        print(f"Endorsement Reason: {reason}")

        # --- Change Basis of SI ---
        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Change Basis of SI").click()

        # --- Sum Insured Selection ---
        page.get_by_role("combobox", name="Plan Sum Insured").click()
        page.get_by_role("option", name="250,000").click()

# --- Extend Period of Insurance ---
def extend_poi(page):
        # --- Expiry Date of Policy ---
        expiry_date = page.locator(".col-md-3").filter(has_text="Expiry Date").locator("span.ng-star-inserted").first.inner_text()
        print("Expiry Date:", expiry_date)

        # ---- Endorsement Reason ----
        try:
            reason = "Extend Period of Insurance"
            page.get_by_text(reason).click()
        except:
            reason = "Extend Expiry Date for Roadtax"
            page.get_by_text(reason).click()

        print(f"Endorsement Reason: {reason}")

        # --- Extending Period of Insurance ---
        expiry_date = datetime.strptime(expiry_date, "%d/%m/%Y")

        # --- Add 2 months manually ---
        month = expiry_date.month + 2
        year = expiry_date.year + (month - 1) // 12
        month = month % 12 or 12
        new_date = expiry_date.replace(year=year, month=month)
        
        # ---- Open Calendar ----
        page.locator("mat-form-field").filter(has_text="Expiry Date").get_by_label("Open calendar").click()

        # --- Extended Date ----
        while True:
            header = page.locator("button.mat-calendar-period-button").inner_text()
            current = datetime.strptime(header.strip().title(), "%b %Y")
            if current.year == new_date.year and current.month == new_date.month:
                break
            page.locator("button.mat-calendar-next-button").click()
            page.wait_for_timeout(300)

        # --- Click the correct day ---
        page.locator(f"button.mat-calendar-body-cell[aria-label='{new_date.strftime('%B')} {new_date.day}, {new_date.year}']").click()
        print("Extended period of Insurance till:", new_date.strftime("%d/%m/%Y"))

        page.get_by_role("button", name="Validate Owner ID # & NCD%").click()

# --- Endorsement Release ---
def endo_release(page):
        # --- Declaration Statement ---
        page.get_by_text("I/We declare that the Policy Information Provided").click()

        # --- Save & Next ----
        page.get_by_role("button", name="Save").click()
        page.get_by_role("button", name="Next").click()

        # --- Endo Quote Reference Number ---
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        endo_quote = quote_text.strip()
        print("Endorsement Quote Number:", endo_quote)

        # --- Evidence Screen ---
        page.get_by_role("button", name="Proceed").click()

        # --- Submit for Processing ---
        page.get_by_role("button", name="Submit for Processing").click()

        print("Endorsement performed successfully")

        return endo_quote

# --- Save Endorsement Quote Reference to Excel ----
def save_excel(policy_number, endo_quote):
    # ====== Save to Excel ======
        BASE_DIR = os.path.join(os.path.dirname(__file__), "..")  # D:\Automation\pages
        file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active

            matched_row = None
            for row in ws.iter_rows(min_row=2):
                cell_policy = row[6]  # Column G (0-indexed = 6) = Policy Number
                if cell_policy.value and str(cell_policy.value).strip() == str(policy_number).strip():
                    matched_row = cell_policy.row
                    break

            if matched_row:

                ws.cell(row=matched_row, column=11).value = endo_quote

                wb.save(file_path)

                print(f"Endo Quote Reference '{endo_quote}' written into Excel, Row {matched_row} (Policy: {policy_number})")
            else:
                print(f"Policy Number '{policy_number}' not found in Excel")