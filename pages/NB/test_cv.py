import sys
import os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from base_login import incep_date, issue_policy, login, navigation, cv_moto
from excel_utils import get_vehicle_data
from datetime import datetime
from nstp_flow import nstp_flow
from openpyxl import Workbook, load_workbook
from test_mail import send_email

# ---- Path References ----
BASE_DIR = os.path.join(os.path.dirname(__file__), "..")               # D:\Automation\pages
DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), "downloads")  # D:\Automation\pages\NB\downloads

def test_cv_motor(page):

    try:
        print("====================== Issuance of CV policy ==================")
        page.wait_for_load_state()
        login(page)
        navigation(page)
        cv_moto(page)

        # ========= FIRST SCREEN ===========

        vehicle_data = get_vehicle_data("CV")
        page.get_by_role("textbox").first.fill(vehicle_data["vehicle_reg_no"])
        print("Registration Number:", vehicle_data["vehicle_reg_no"])

        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Johor").click()

        page.get_by_role("button", name="search Vehicle Search").click()

        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Commercial Vehicle").click()

        # ---- VEHICLE USE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="C permit").click()

        page.locator("mat-form-field").filter(has_text="Engine # *").locator("#engineNo").fill("63547")
        page.locator("mat-form-field").filter(has_text="Chassis # *").locator("#chassisNo").fill("34576657")

        # ---- MAKE & MODEL ----
        page.locator("mat-form-field", has_text="Make").click()
        page.get_by_role("option", name="VOLVO").click()

        page.locator("mat-form-field", has_text="Model").click()
        page.get_by_role("option", name="F16").click()

        # ---- Year of Manufacture ----
        page.locator("mat-form-field", has_text="Year of Manufacture").click()
        page.get_by_role("option", name="2015").click()
        page.wait_for_timeout(2000)

        # ---- Vehicle Age (to determine coverage type) ---- 
        vehicle_age_locator = page.locator("mat-form-field").filter(has_text="Vehicle Age").locator("#vehicleAge")
        vehicle_age_text = vehicle_age_locator.input_value().strip()

        vehicle_age = int(vehicle_age_text)
        print(f"Vehicle Age: {vehicle_age} years")

        # ---- VARIANT ----
        page.locator(".mat-select-placeholder.mat-select-min-line.ng-tns-c176-76").click()
        page.get_by_role("option", name="NA").click()

        # ---- Seating Capacity ----
        page.locator("mat-form-field").filter(has_text="Seating Capacity *").locator("#seatCapacity").fill("5")
        # ---- Carrying Capacity ----
        page.locator("mat-form-field").filter(has_text="Carrying Capacity *").locator("#carryingCapacity").fill("20")

        page.locator(".mat-select-placeholder.mat-select-min-line.ng-tns-c176-84").click()
        page.get_by_role("option", name="kg").click()

        page.locator(".mat-select-placeholder.mat-select-min-line.ng-tns-c176-86").click()
        page.get_by_role("option", name="Beverages Bottles").click()

        # ---- Save Vehicle Info Button ----
        page.get_by_role("button", name="Save Vehicle Info").click()


        # ========== SECOND SCREEN ==========

        # ---- Coverage Type ----
        page.locator("#mat-select-value-31").click()
        if vehicle_age >= 20:
            page.get_by_role("option", name="TP, Fire & Theft").click()
        else:
            page.get_by_role("option", name="Comprehensive").click()

        selected_coverage = page.locator("#mat-select-value-31").inner_text().strip()
        print("Selected Coverage type: ", selected_coverage)

        # ---- COVERAGE DATE -----
        incep_date(page)

        # ----- SUM INSURED ----
        page.get_by_role("region", name="Coverage").locator("input[type=\"text\"]").click()
        page.get_by_role("region", name="Coverage").locator("input[type=\"text\"]").fill("20000")

        # ---- BUSINESS REGISTRATION NUMBER ----
        page.locator("mat-form-field").filter(has_text="Business Registration # *").locator("#id").fill(vehicle_data["mykad"])
        print("MyKad Number:", vehicle_data["mykad"])

        # ---- NAME AS PER ID / LEGAL NAME ----
        page.locator("dx-input").filter(has_text="* Name as per ID / Legal Name").locator("#legalName").fill("CV C Permit")
        page.get_by_role("button", name="search Validate Owner as per").click()

        # ---- NCD value ----
        page.wait_for_timeout(7000)
        ncd_value = page.locator("#currentNCD input.mat-input-element").input_value()
        print("NCD Value:", ncd_value)
        
        # --- SAVE & NEXT BUTTON ----
        page.get_by_role("button", name="Save & Next").click()
        print("Registration Number is Triggered to ISM")

        page.wait_for_timeout(5000)
        
        # ========== THIRD SCREEN ==== PH Details ======
        
        # ---- CHECK IF YES BUTTON EXISTS AND IS ENABLED ----
        try:
            yes_button = page.get_by_role("button", name="Yes").first
            yes_button.wait_for(state="visible", timeout=5000)
            yes_button.click()
            page.wait_for_timeout(1000)
            print("Yes button clicked")
        except:
            print("Yes button not visible, skipping")

        # ---- CHECK IF ADDRESS ALREADY EXISTS ----
        try:
            add_button = page.locator("button[name='Add'], button:has-text('Add')").first
            add_button.wait_for(state="visible", timeout=5000)
            add_button.click()
            print("Add button clicked")
            page.wait_for_timeout(2000)
        except:
            print("Add button not visible, skipping")

        # ---- STATE ---- (runs for both cases)
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Johor").click()
        page.wait_for_timeout(2000)

        # ---- PINCODE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="81100").click()
        page.wait_for_timeout(1000)

        # ---- STREET ADDRESS ----
        page.get_by_role("combobox", name="Address Line").click()
        page.get_by_role("option", name="Taman Desa Harmoni", exact=True).click()
        page.wait_for_timeout(1000)

        # ---- SAVE BUTTON (if address is added) ----
        address_save = page.locator("button#save")
        if address_save.is_visible():
            address_save.click()

        # ---- Declaration Statements ----
        page.get_by_text("We respect your privacy and").click()
        page.get_by_text("I hereby confirm that I have").click()

        # ---- Get Quote Number ----
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        quote_number = quote_text.strip()
        print("Quote Number:", quote_number)

        # ====== NSTP FLOW FUNCTION CALL ======
        generate_quote_btn = nstp_flow(page, quote_number, vehicle_type="cv")

        # ---- Generate Quote Flow ----
        if generate_quote_btn.is_visible():
            generate_quote_btn.click()
            print("STP process, clicked on Generate Quote button")

            with page.expect_download() as download_info:
                page.get_by_role("button", name="Submit").click()
            download_info.value.save_as(os.path.join(DOWNLOADS_DIR, "CV_quote.pdf"))

        # ==== Issue Policy function ====
        policy_number = issue_policy(page)
        
        # ---- Download the policy schedule ----
        page.get_by_role("button", name="Download & e-mail Policy").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as(os.path.join(DOWNLOADS_DIR, "CV_policy.pdf"))

        print("Policy is Issued and Schedule letter downloaded successfully.")

        
        # --------- SAVE TO EXCEL ---------
        
        file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

        # Load or create workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # ---- Find next empty row based on Column C (NV/RV) ----
        row = 2
        while ws.cell(row=row, column=3).value:
            row += 1

        # ---- Serial Number in Column A ----
        if row == 2:
            serial_no = 1
        else:
            prev_serial = ws.cell(row=row - 1, column=1).value
            serial_no = (prev_serial or 0) + 1

        # ---- Policy Type ----
        registration = "RV"
        policy_type = "CV"

        inception_date_excel = datetime.today().strftime("%d-%m-%Y")

        # ---- Write data ----
        ws.cell(row=row, column=1).value = serial_no                # Column A - Serial Number
        ws.cell(row=row, column=3).value = registration             # Column C - NV/RV
        ws.cell(row=row, column=4).value = policy_type              # Column D - Policy Type
        ws.cell(row=row, column=5).value = selected_coverage        # Column E - Coverage Type
        ws.cell(row=row, column=6).value = quote_number             # Column F - Quote Number
        ws.cell(row=row, column=7).value = policy_number            # Column G - Policy Number
        ws.cell(row=row, column=8).value = inception_date_excel     # Column H - Inception Date

        # ---- Auto-fill Column I & J from previous row (like Ctrl+D) ----
        if row > 2:
            prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
            prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
            prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

            if prev_col_b:
                ws.cell(row=row, column=2).value = prev_col_b   # Column B
            if prev_col_i:
                ws.cell(row=row, column=9).value = prev_col_i   # Column I
            if prev_col_j:
                ws.cell(row=row, column=10).value = prev_col_j  # Column J

        # ---- Save file ----
        wb.save(file_path)

        print("In Excel, Quote and Policy numbers captured successfully")

        # -------- SEND EMAIL ---------
        try:
            send_email()
        except Exception as e:
            print("Email failed:", e)


    finally:
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()
        page.wait_for_timeout(15000)