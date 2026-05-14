import os
import traceback
from base_login import login, navigation, cv_moto, manager_approval
from excel_utils import get_vehicle_data
from datetime import datetime
from openpyxl import Workbook, load_workbook
from test_mail import send_email


def test_cv_motor(page):

    try:
        page.wait_for_load_state
        login(page)
        navigation(page)
        cv_moto(page)

        # ========= FIRST SCREEN ===========

        vehicle_data = get_vehicle_data("CV")
        page.get_by_role("textbox").first.fill(vehicle_data["vehicle_reg_no"])
        print("Registration Number:", vehicle_data["vehicle_reg_no"])

        page.locator(".mat-select-placeholder").click()
        page.get_by_role("option", name="Johor").click()

        page.get_by_role("button", name="search Vehicle Search").click()

        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Commercial Vehicle").click()

        # ---- VEHICLE USE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="C permit").click()

        page.locator("mat-form-field").filter(has_text="Engine # *").locator("#engineNo").fill("63547")
        page.locator("mat-form-field").filter(has_text="Chassis # *").locator("#chassisNo").fill("34576657")

        # ---- MAKE & MODEL ----
        page.locator("mat-form-field", has_text="Make").click()
        page.get_by_role("option", name="VOLVO").click()

        page.locator("mat-form-field", has_text="Model").click()
        page.get_by_role("option", name="F16").click()

        # ---- Year of Manufacture ----
        page.locator(".mat-select-placeholder.mat-select-min-line.ng-tns-c176-74").click()
        page.get_by_role("option", name="2005").click()
        page.wait_for_timeout(2000)

        # ---- Vehicle Age (to determine coverage type) ---- 
        vehicle_age_locator = page.locator("mat-form-field").filter(has_text="Vehicle Age").locator("#vehicleAge")
        vehicle_age_text = vehicle_age_locator.input_value().strip()

        vehicle_age = int(vehicle_age_text)
        print(f"Vehicle Age: {vehicle_age} years")

        # ---- VARIANT ----
        page.locator(".mat-select-placeholder.mat-select-min-line.ng-tns-c176-76").click()
        page.get_by_role("option", name="NA").click()

        # ---- Seating Capacity ----
        page.locator("mat-form-field").filter(has_text="Seating Capacity *").locator("#seatCapacity").fill("5")
        # ---- Carrying Capacity ----
        page.locator("mat-form-field").filter(has_text="Carrying Capacity *").locator("#carryingCapacity").fill("1")

        page.locator(".mat-select-placeholder.mat-select-min-line.ng-tns-c176-84").click()
        page.get_by_role("option", name="kg").click()
        #page.get_by_role("option", name="Tonnes").click()

        page.locator(".mat-select-placeholder.mat-select-min-line.ng-tns-c176-86").click()
        page.get_by_role("option", name="Beverages Bottles").click()

        # ---- Save Vehicle Info Button ----
        page.get_by_role("button", name="Save Vehicle Info").click()


        # ========== SECOND SCREEN ==========

        # ---- Coverage Type ----
        page.locator("#mat-select-value-31").click()
        if vehicle_age >= 20:
            page.get_by_role("option", name="TP, Fire & Theft").click()
        else:
            page.get_by_role("option", name="Comprehensive").click()

        selected_coverage = page.locator("#mat-select-value-31").inner_text().strip()
        print("Selected Coverage type: ", selected_coverage)

        # ---- INCEPTION DATE (Today's date) ----
        today = datetime.today()
        aria_date = today.strftime("%B %d, %Y").replace(" 0", " ")

        page.locator("mat-form-field") \
            .filter(has_text="Inception Date * event") \
            .get_by_label("Open calendar") \
            .click()

        page.get_by_role("gridcell", name=aria_date).click()

        # ----- SUM INSURED ----
        page.get_by_role("region", name="Coverage").locator("input[type=\"text\"]").click()
        page.get_by_role("region", name="Coverage").locator("input[type=\"text\"]").fill("20000")

        # ---- BUSINESS REGISTRATION NUMBER ----
        page.locator("mat-form-field").filter(has_text="Business Registration # *").locator("#id").fill(vehicle_data["mykad"])
        print("MyKad Number:", vehicle_data["mykad"])

        # ---- NAME AS PER ID / LEGAL NAME ----
        page.locator("dx-input").filter(has_text="* Name as per ID / Legal Name").locator("#legalName").fill("CV C Permit")
        page.get_by_role("button", name="search Validate Owner as per").click()

        # ---- NCD value ----
        page.wait_for_timeout(7000)
        ncd_value = page.locator("#currentNCD input.mat-input-element").input_value()
        print("NCD Value:", ncd_value)
        
        # --- SAVE & NEXT BUTTON ----
        page.get_by_role("button", name="Save & Next").click()
        print("Registration Number is Triggered to ISM")

        page.wait_for_timeout(5000)
        
        # ========== THIRD SCREEN ==== PH Details ======

        # ---- CHECK IF YES BUTTON EXISTS AND IS ENABLED ----
        try:
            yes_button = page.get_by_role("button", name="Yes").first
            yes_button.wait_for(state="visible", timeout=5000)
            yes_button.click()
            page.wait_for_timeout(1000)
            print("Yes button clicked")
        except:
            print("Yes button not visible, skipping")

        # ---- CHECK IF ADDRESS ALREADY EXISTS ----
        try:
            add_button = page.locator("button[name='Add'], button:has-text('Add')").first
            add_button.wait_for(state="visible", timeout=5000)
            add_button.click()
            print("Add button clicked")
            page.wait_for_timeout(2000)
        except:
            print("Add button not visible, skipping")

        # ---- STATE ---- (runs for both cases)
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="Johor").click()
        page.wait_for_timeout(2000)

        # ---- PINCODE ----
        page.locator(".mat-select-placeholder").first.click()
        page.get_by_role("option", name="81100").click()
        page.wait_for_timeout(1000)

        # ---- STREET ADDRESS ----
        page.get_by_role("combobox", name="Address Line").click()
        page.get_by_role("option", name="Taman Desa Harmoni", exact=True).click()
        page.wait_for_timeout(1000)

        # ---- SAVE BUTTON (if address is added) ----
        address_save = page.locator("button#save")
        if address_save.is_visible():
            address_save.click()

        # ---- DECLARATION STATEMENTS ----
        page.get_by_text("We respect your privacy and").click()
        page.get_by_text("I hereby confirm that I have").click()

        # ---- Get Quote Number ----
        quote_text = page.locator("text=Quote Reference #").locator("xpath=following-sibling::*").inner_text()
        quote_number = quote_text.strip()
        print("Quote Number:", quote_number)

        # ===== THIRD SCREEN FLOW =====

        submit_approval_btn = page.get_by_role("button", name="Submit for TPM Staff Approval")
        generate_quote_btn = page.get_by_role("button", name="Generate Quote")

        # ---- Approval Flow ----
        if submit_approval_btn.is_visible():
            submit_approval_btn.click()
            page.wait_for_timeout(17000)

            # ---- Incognito Session ----
            browser = page.context.browser
            manager_page = browser.new_context().new_page()
            manager_page.goto(f"https://agent-uat.tuneinsurance.com/#/qms/quote/motor/rcv/cover-details?edit=true&quoteNr={quote_number}")

            manager_approval(manager_page)

            page.wait_for_timeout(10000)
            page.get_by_role("button", name="Back").click()
            page.wait_for_timeout(3000)

        # ---- Generate Quote Flow ----
        if generate_quote_btn.is_visible():
            generate_quote_btn.click()

            with page.expect_download() as download_info:
                page.get_by_role("button", name="Submit").click()
            download_info.value.save_as("downloads/CV_quote.pdf")

            page.get_by_role( "button", name="Proceed to Policy Issuance").click()

            # ----- ISSUE POLICY -----
            page.get_by_role("button", name="Issue Policy").click()

        else:
            print("Generate Quote button not visible")

        page.wait_for_timeout(25000)

        page.reload()

        page.wait_for_timeout(20000)

        # ---- Printing the policy number ---
        policy_locator = page.locator("text=Policy #:")
        policy_locator.wait_for()
        policy_text = policy_locator.text_content()
        policy_number = policy_text.replace("Policy #:", "").strip()
        print("Policy Number:", policy_number)

        # ---- Download the policy schedule ----
        page.get_by_role("button", name="Download & e-mail Policy").click()
        page.wait_for_timeout(5000)

        with page.expect_download() as download_info:
            page.get_by_role("button", name="Submit").click()
        download = download_info.value
        download.save_as("downloads/CV_policy.pdf")


        # --------- SAVE TO EXCEL ---------
        
        file_path = "UATStability.xlsx"

        # Load or create workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active

        # ---- Find next empty row based on Column C (NV/RV) ----
        row = 2
        while ws.cell(row=row, column=3).value:
            row += 1

        # ---- Serial Number in Column A ----
        if row == 2:
            serial_no = 1
        else:
            prev_serial = ws.cell(row=row - 1, column=1).value
            serial_no = (prev_serial or 0) + 1

        # ---- Policy Type ----
        registration = "RV"
        policy_type = "CV"

        inception_date_excel = today.strftime("%d-%m-%Y")

        # ---- Write data ----
        ws.cell(row=row, column=1).value = serial_no                # Column A - Serial Number
        ws.cell(row=row, column=3).value = registration             # Column C - NV/RV
        ws.cell(row=row, column=4).value = policy_type              # Column D - Policy Type
        ws.cell(row=row, column=5).value = selected_coverage        # Column E - Coverage Type
        ws.cell(row=row, column=6).value = quote_number             # Column F - Quote Number
        ws.cell(row=row, column=7).value = policy_number            # Column G - Policy Number
        ws.cell(row=row, column=8).value = inception_date_excel     # Column H - Inception Date

        # ---- Auto-fill Column I & J from previous row (like Ctrl+D) ----
        if row > 2:
            prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
            prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
            prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

            if prev_col_b:
                ws.cell(row=row, column=2).value = prev_col_b   # Column B
            if prev_col_i:
                ws.cell(row=row, column=9).value = prev_col_i   # Column I
            if prev_col_j:
                ws.cell(row=row, column=10).value = prev_col_j  # Column J


        # ---- Save file ----
        wb.save(file_path)

        print("In Excel, Quote and Policy numbers captured successfully")

        # -------- SEND EMAIL ---------
        try:
            send_email()
        except Exception as e:
            print("Email failed:", e)

    except Exception as e:
    # ADD THIS - captures the real error
        print("Test failed:", traceback.format_exc())

    finally:
        page.get_by_text("playwright", exact=True).click()
        page.get_by_text("Sign Out", exact=True).click()
        page.wait_for_timeout(15000)