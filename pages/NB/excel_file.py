import os
from datetime import datetime
from openpyxl import Workbook, load_workbook

# ---- Path References ----
BASE_DIR = os.path.join(os.path.dirname(__file__), "..")               # D:\Automation\pages
DOWNLOADS_DIR = os.path.join(os.path.dirname(__file__), "downloads")  # D:\Automation\pages\NB\downloads

# --- PA excel file ------
def pa_excel(selected_title, quote_number, policy_number):

    # ------- SAVE TO EXCEL -------
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

    # Load or create workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # ---- Find next empty row based on Column C (NV/RV) ----
    row = 2
    while ws.cell(row=row, column=3).value:
        row += 1

    # ---- Serial Number in Column A ----
    if row == 2:
        serial_no = 1
    else:
        prev_serial = ws.cell(row=row - 1, column=1).value
        serial_no = (prev_serial or 0) + 1

    # ---- Policy Type ----
    registration = "NV"
    policy_type = "PA"
    inception_date_excel = datetime.today().strftime("%d-%m-%Y")

    # ---- Write data ----
    ws.cell(row=row, column=1).value = serial_no        # Column A - Serial Number
    ws.cell(row=row, column=3).value = registration     # Column C - NV/RV
    ws.cell(row=row, column=4).value = policy_type      # Column D - Policy Type
    ws.cell(row=row, column=5).value = selected_title   # Column E - Coverage Type
    ws.cell(row=row, column=6).value = quote_number     # Column F - Quote Number
    ws.cell(row=row, column=7).value = policy_number    # Column G - Policy Number
    ws.cell(row=row, column=8).value = inception_date_excel

    # ---- Auto-fill Column B, I & J from previous row (like Ctrl+D) ----
    if row > 2:
        prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
        prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
        prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

        if prev_col_b:
            ws.cell(row=row, column=2).value = prev_col_b   # Column B
        if prev_col_i:
            ws.cell(row=row, column=9).value = prev_col_i   # Column I
        if prev_col_j:
            ws.cell(row=row, column=10).value = prev_col_j  # Column J

    # ---- Save file ----
    wb.save(file_path)

    print("In Excel, Quote and Policy numbers captured successfully")

# --- MC excel file ------
def mc_excel(selected_coverage, quote_number, policy_number):
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

    # Load or create workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # ---- Find next empty row based on Column C (NV/RV) ----
    row = 2  # start after header
    while ws.cell(row=row, column=3).value:
        row += 1

    # ---- Serial Number in Column A ----
    if row == 2:
        serial_no = 1
    else:
        prev_serial = ws.cell(row=row - 1, column=1).value
        serial_no = (prev_serial or 0) + 1

    # ---- Policy Type ----
    registration = "RV"
    policy_type = "MC"

    inception_date_excel = datetime.today().strftime("%d-%m-%Y")

    # ---- Write data ----
    ws.cell(row=row, column=1).value = serial_no             # Column A - Serial Number
    ws.cell(row=row, column=3).value = registration          # Column C - NV/RV
    ws.cell(row=row, column=4).value = policy_type           # Column D - Policy Type
    ws.cell(row=row, column=5).value = selected_coverage     # Column E - Coverage Type
    ws.cell(row=row, column=6).value = quote_number          # Column F - Quote Number
    ws.cell(row=row, column=7).value = policy_number         # Column G - Policy Number
    ws.cell(row=row, column=8).value = inception_date_excel  # Column H

    # ---- Auto-fill Column I & J from previous row (like Ctrl+D) ----
    if row > 2:
        prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
        prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
        prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

        if prev_col_b:
            ws.cell(row=row, column=2).value = prev_col_b   # Column B
        if prev_col_i:
            ws.cell(row=row, column=9).value = prev_col_i   # Column I
        if prev_col_j:
            ws.cell(row=row, column=10).value = prev_col_j  # Column J

    # ---- Save file ----
    wb.save(file_path)

    print("In Excel, Quote and Policy numbers captured successfully")

# --- PC excel file ------
def pc_excel(selected_coverage, quote_number, policy_number):
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

    # Load or create workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # ---- Find next empty row based on Column C (NV/RV) ----
    row = 2  
    while ws.cell(row=row, column=3).value:
        row += 1
    
    # ---- Serial Number in Column A ----
    if row == 2:
        serial_no = 1
    else:
        prev_serial = ws.cell(row=row - 1, column=1).value
        serial_no = (prev_serial or 0) + 1

    # ---- Policy Type ----
    registration = "RV"
    policy_type = "PC"

    inception_date_excel = datetime.today().strftime("%d-%m-%Y")

    # ---- Write data ----
    ws.cell(row=row, column=1).value = serial_no         # Column A - Serial Number
    ws.cell(row=row, column=3).value = registration      # Column C - NV/RV
    ws.cell(row=row, column=4).value = policy_type       # Column D - Policy Type
    ws.cell(row=row, column=5).value = selected_coverage # Column E - Coverage Type
    ws.cell(row=row, column=6).value = quote_number      # Column F - Quote Number
    ws.cell(row=row, column=7).value = policy_number     # Column G - Policy Number
    ws.cell(row=row, column=8).value = inception_date_excel       

    # ---- Auto-fill Column I & J from previous row  ----
    if row > 2: 
        prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
        prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
        prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

        if prev_col_b:
            ws.cell(row=row, column=2).value = prev_col_b   # Column B
        if prev_col_i:
            ws.cell(row=row, column=9).value = prev_col_i   # Column I
        if prev_col_j:
            ws.cell(row=row, column=10).value = prev_col_j  # Column J

    # ---- Save file ----
    wb.save(file_path)

    print("In Excel, Quote and Policy numbers captured successfully")

# --- CV excel file ------
def cv_excel(selected_coverage, quote_number, policy_number):
    
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

    # Load or create workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # ---- Find next empty row based on Column C (NV/RV) ----
    row = 2
    while ws.cell(row=row, column=3).value:
        row += 1

    # ---- Serial Number in Column A ----
    if row == 2:
        serial_no = 1
    else:
        prev_serial = ws.cell(row=row - 1, column=1).value
        serial_no = (prev_serial or 0) + 1

    # ---- Policy Type ----
    registration = "RV"
    policy_type = "CV"

    inception_date_excel = datetime.today().strftime("%d-%m-%Y")

    # ---- Write data ----
    ws.cell(row=row, column=1).value = serial_no                # Column A - Serial Number
    ws.cell(row=row, column=3).value = registration             # Column C - NV/RV
    ws.cell(row=row, column=4).value = policy_type              # Column D - Policy Type
    ws.cell(row=row, column=5).value = selected_coverage        # Column E - Coverage Type
    ws.cell(row=row, column=6).value = quote_number             # Column F - Quote Number
    ws.cell(row=row, column=7).value = policy_number            # Column G - Policy Number
    ws.cell(row=row, column=8).value = inception_date_excel     # Column H - Inception Date

    # ---- Auto-fill Column I & J from previous row (like Ctrl+D) ----
    if row > 2:
        prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
        prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
        prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

        if prev_col_b:
            ws.cell(row=row, column=2).value = prev_col_b   # Column B
        if prev_col_i:
            ws.cell(row=row, column=9).value = prev_col_i   # Column I
        if prev_col_j:
            ws.cell(row=row, column=10).value = prev_col_j  # Column J

    # ---- Save file ----
    wb.save(file_path)

    print("In Excel, Quote and Policy numbers captured successfully")

def dental(quote_number, policy_number):
    # ------- SAVE TO EXCEL -------
    file_path = os.path.join(BASE_DIR, "UATStability.xlsx")

    # Load or create workbook
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    # ---- Find next empty row based on Column C (NV/RV) ----
    row = 2
    while ws.cell(row=row, column=3).value:
        row += 1

    # ---- Serial Number in Column A ----
    if row == 2:
        serial_no = 1
    else:
        prev_serial = ws.cell(row=row - 1, column=1).value
        serial_no = (prev_serial or 0) + 1

    # ---- Policy Type ----
    registration = "NV"
    policy_type = "Dental"
    selected_title = "Dental Shield"
    inception_date_excel = datetime.today().strftime("%d-%m-%Y")

    # ---- Write data ----
    ws.cell(row=row, column=1).value = serial_no        # Column A - Serial Number
    ws.cell(row=row, column=3).value = registration     # Column C - NV/RV
    ws.cell(row=row, column=4).value = policy_type      # Column D - Policy Type
    ws.cell(row=row, column=5).value = selected_title   # Column E - Coverage Type
    ws.cell(row=row, column=6).value = quote_number     # Column F - Quote Number
    ws.cell(row=row, column=7).value = policy_number    # Column G - Policy Number
    ws.cell(row=row, column=8).value = inception_date_excel

    # ---- Auto-fill Column B, I & J from previous row (like Ctrl+D) ----
    if row > 2:
        prev_col_b = ws.cell(row=row - 1, column=2).value   # Column B
        prev_col_i = ws.cell(row=row - 1, column=9).value   # Column I
        prev_col_j = ws.cell(row=row - 1, column=10).value  # Column J

        if prev_col_b:
            ws.cell(row=row, column=2).value = prev_col_b   # Column B
        if prev_col_i:
            ws.cell(row=row, column=9).value = prev_col_i   # Column I
        if prev_col_j:
            ws.cell(row=row, column=10).value = prev_col_j  # Column J

    # ---- Save file ----
    wb.save(file_path)

    print("In Excel, Quote and Policy numbers captured successfully")
