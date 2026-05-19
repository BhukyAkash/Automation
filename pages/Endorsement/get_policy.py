import openpyxl
from datetime import datetime, date

# ===== Function to get policy number for a given product from Excel =======
def get_policy_number(product):
    # ---- Open the UAT Stability Excel file ----
    wb = openpyxl.load_workbook(r"D:\Automation\pages\UATStability.xlsx")
    ws = wb.active

    # ---- Read header row to find column positions by name ----
    headers = [cell.value for cell in ws[1]]
    motor_col = headers.index("Motor Type")
    policy_col = headers.index("Policy Number")
    date_col = headers.index("Created On")

    # ---- Get today's date to match against Created On ----
    today = date.today()

    # ---- Map CLI --product value to actual Motor Type values in Excel ----
    motor_types = {
        "motor": ["MC", "PC", "CV"],
        "pa": ["PA"],
        "mc": ["MC"],
        "pc": ["PC"],
        "cv": ["CV"],
    }
    # Get the list of valid Motor Types for the given product
    valid_types = motor_types.get(product.lower(), [product.upper()])

    # ---- Loop through each data row in Excel  ----
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_motor = str(row[motor_col]).strip().upper() if row[motor_col] else ""
        row_policy = row[policy_col]
        row_date = row[date_col]

        # ---- Skip row if Motor Type doesn't match product or Policy Number is empty ----
        if row_motor not in valid_types or not row_policy:
            continue

        # ---- Normalize date to a date object for comparison ----
        # Excel stores dates as datetime objects
        if isinstance(row_date, datetime):
            row_date = row_date.date()

        # If date is stored as a string, parse it
        elif isinstance(row_date, str):
            for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    row_date = datetime.strptime(row_date, fmt).date()
                    break  # Stop trying formats once one works
                except ValueError:
                    continue
            else:
                continue

        # ---- Return policy only if it was created today ----
        if row_date == today:
            return str(row_policy).strip(), row_motor, row_date

    # ---- No matching policy found for today ----
    return None, None, None